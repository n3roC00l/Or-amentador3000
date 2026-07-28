"""
Gera o "Mapa de Cotação" em Excel a partir do banco — mesmo layout visual
do modelo atual (FILAMENTOS_..._Mapa_de_Cotação.xlsx), mas com a coluna
"ORÇAMENTO MAIS ECONÔMICO" calculada de verdade via MIN()/IF(), em vez de
fixada no primeiro fornecedor como no arquivo original.

Uso:
    python export_excel.py [objeto] [saida.xlsx]

Fonte dos dados: sempre a leitura mais recente de cada (item, fornecedor)
em `cotacoes`. Item sem nenhuma leitura 'ok'/'suspeito' fica em branco na
planilha — nunca vira 0 silencioso.
"""
import sys
import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from db import conectar

FONTE = "Arial"
NEGRITO = Font(name=FONTE, bold=True)
NORMAL = Font(name=FONTE)
FUNDO_DESTAQUE = PatternFill("solid", fgColor="FFF2CC")
CENTRO = Alignment(horizontal="center", vertical="center")
MOEDA = "R$#,##0.00"


def _dados():
    conn = conectar()
    fornecedores = conn.execute("SELECT * FROM fornecedores ORDER BY id").fetchall()
    itens = conn.execute("SELECT * FROM itens ORDER BY ordem").fetchall()

    cotacoes = {}  # (item_id, fornecedor_id) -> row mais recente
    linhas = conn.execute("""
        SELECT c.* FROM cotacoes c
        JOIN (
            SELECT item_id, fornecedor_id, MAX(coletado_em) AS ultima
            FROM cotacoes GROUP BY item_id, fornecedor_id
        ) r ON c.item_id = r.item_id AND c.fornecedor_id = r.fornecedor_id
           AND c.coletado_em = r.ultima
    """).fetchall()
    for linha in linhas:
        cotacoes[(linha["item_id"], linha["fornecedor_id"])] = linha

    conn.close()
    return fornecedores, itens, cotacoes


def gerar(objeto: str = "Filamentos 3D", caminho_saida: str = "mapa_de_cotacao.xlsx"):
    fornecedores, itens, cotacoes = _dados()
    n = len(fornecedores)
    m = len(itens)
    if n == 0 or m == 0:
        raise SystemExit("Cadastre fornecedores e itens antes de exportar (rode seed.py).")

    wb = Workbook()
    ws = wb.active
    ws.title = "Mapa de Cotação"

    # --- coordenadas dos blocos de fornecedor: (col_unit, col_total, nome_cel, frete_cel)
    blocos = []
    for i in range(n):
        col_unit = 5 + 2 * i          # E, G, I, ...
        col_total = col_unit + 1      # F, H, J, ...
        blocos.append({"unit": col_unit, "total": col_total, "fornecedor": fornecedores[i]})
    col_min_unit = 5 + 2 * n
    col_min_total = col_min_unit + 1
    col_empresa = col_min_total + 1

    def C(col, row):
        return f"{get_column_letter(col)}{row}"

    # --- cabeçalho -----------------------------------------------------
    ws["A1"] = "MAPA DE COTAÇÃO DE PREÇOS"
    ws["A1"].font = Font(name=FONTE, bold=True, size=14)

    ws["A2"] = "Objeto:"
    ws["A2"].font = NEGRITO
    ws["C2"] = objeto

    ws["A3"] = "DATA DA COTAÇÃO:"
    ws["A3"].font = NEGRITO
    ws["D3"] = datetime.date.today()
    ws["D3"].number_format = "DD/MM/YYYY"

    row_fornecedores_header = 4
    ws[C(1, row_fornecedores_header)] = "N.º"
    ws[C(2, row_fornecedores_header)] = "F O R N E C E D O R E S"
    ws[C(6, row_fornecedores_header)] = "PRAZO INICIO"
    ws[C(7, row_fornecedores_header)] = "CONDIÇÕES PAGAMENTO"
    ws[C(9, row_fornecedores_header)] = "PRAZO PARA ENTREGA"
    for col in (1, 2, 6, 7, 9):
        ws[C(col, row_fornecedores_header)].font = NEGRITO

    linha_nome_fornecedor = {}  # fornecedor_id -> linha onde o nome está (p/ fórmulas)
    for i, bloco in enumerate(blocos):
        f = bloco["fornecedor"]
        r = 5 + i
        linha_nome_fornecedor[f["id"]] = r
        ws[C(1, r)] = i + 1
        ws[C(2, r)] = f["nome"]
        ws[C(7, r)] = f["condicao_pagamento"]
        ws[C(9, r)] = f["prazo_entrega"]
        ws[C(11, r)] = f["site"]

    # --- tabela de itens -------------------------------------------------
    linha_hdr1 = 5 + n + 2
    linha_hdr2 = linha_hdr1 + 1
    linha_item0 = linha_hdr2 + 1

    ws[C(1, linha_hdr1)] = "ITEM"
    ws[C(2, linha_hdr1)] = "QUANT."
    ws[C(3, linha_hdr1)] = "UND"
    ws[C(4, linha_hdr1)] = "ESPECIFICAÇÕES DOS PRODUTOS/SERVIÇOS"
    for bloco in blocos:
        ws[C(bloco["unit"], linha_hdr1)] = bloco["fornecedor"]["nome"]
    ws[C(col_min_unit, linha_hdr1)] = "ORÇAMENTO MAIS ECONÔMICO"

    for col in (1, 2, 3, 4, col_min_unit):
        ws[C(col, linha_hdr1)].font = NEGRITO
    for bloco in blocos:
        ws[C(bloco["unit"], linha_hdr1)].font = NEGRITO

    for bloco in blocos:
        ws[C(bloco["unit"], linha_hdr2)] = "VALOR UNITÁRIO"
        ws[C(bloco["total"], linha_hdr2)] = "VALOR TOTAL"
        ws[C(bloco["unit"], linha_hdr2)].font = NEGRITO
        ws[C(bloco["total"], linha_hdr2)].font = NEGRITO
    ws[C(col_min_unit, linha_hdr2)] = "VALOR UNITÁRIO"
    ws[C(col_min_total, linha_hdr2)] = "VALOR TOTAL"
    ws[C(col_empresa, linha_hdr2)] = "EMPRESA"
    for col in (col_min_unit, col_min_total, col_empresa):
        ws[C(col, linha_hdr2)].font = NEGRITO

    # --- linhas de item ----------------------------------------------------
    for idx, item in enumerate(itens):
        r = linha_item0 + idx
        ws[C(1, r)] = idx + 1
        ws[C(2, r)] = item["quantidade"]
        ws[C(3, r)] = item["unidade"]
        ws[C(4, r)] = item["especificacao"]

        unit_cells = []
        for bloco in blocos:
            f = bloco["fornecedor"]
            cot = cotacoes.get((item["id"], f["id"]))
            unit_col_letter = get_column_letter(bloco["unit"])
            tem_valor = cot is not None and cot["preco_unitario"] is not None and cot["status"] != "falha"
            if tem_valor:
                ws[C(bloco["unit"], r)] = cot["preco_unitario"]
                ws[C(bloco["unit"], r)].number_format = MOEDA
                if cot["status"] == "suspeito":
                    ws[C(bloco["unit"], r)].fill = PatternFill("solid", fgColor="FCE4D6")
                unit_cells.append(f"{unit_col_letter}{r}")
                ws[C(bloco["total"], r)] = f"={unit_col_letter}{r}*$B{r}"
            # sem leitura válida: célula(s) ficam em branco (não entram no MIN,
            # não viram 0 silencioso)
            ws[C(bloco["total"], r)].number_format = MOEDA

        if unit_cells:
            ws[C(col_min_unit, r)] = f"=MIN({','.join(unit_cells)})"
            ws[C(col_min_total, r)] = f"={get_column_letter(col_min_unit)}{r}*$B{r}"
            # nested IF: casa o valor mínimo com o nome do fornecedor certo
            expressao = '""'
            for bloco in reversed(blocos):
                unit_letter = get_column_letter(bloco["unit"])
                nome_cel = f"$B${linha_nome_fornecedor[bloco['fornecedor']['id']]}"
                expressao = f'IF({unit_letter}{r}={get_column_letter(col_min_unit)}{r},{nome_cel},{expressao})'
            ws[C(col_empresa, r)] = f"={expressao}"
        ws[C(col_min_unit, r)].number_format = MOEDA
        ws[C(col_min_total, r)].number_format = MOEDA
        ws[C(col_min_unit, r)].fill = FUNDO_DESTAQUE
        ws[C(col_min_total, r)].fill = FUNDO_DESTAQUE

    # --- total, frete -----------------------------------------------------
    linha_total = linha_item0 + m
    linha_frete = linha_total + 1
    primeira, ultima = linha_item0, linha_item0 + m - 1

    for bloco in blocos:
        total_letter = get_column_letter(bloco["total"])
        ws[C(bloco["total"], linha_total)] = f"=SUM({total_letter}{primeira}:{total_letter}{ultima})"
        ws[C(bloco["total"], linha_total)].number_format = MOEDA
        ws[C(bloco["unit"], linha_frete)] = bloco["fornecedor"]["frete"]
        ws[C(bloco["unit"], linha_frete)].number_format = MOEDA
        ws[C(bloco["unit"], linha_frete)].font = Font(name=FONTE, color="0000FF")  # input manual

    ws[C(col_min_unit, linha_total)] = "TOTAL"
    ws[C(col_min_unit, linha_total)].font = NEGRITO
    min_total_letter = get_column_letter(col_min_total)
    ws[C(col_min_total, linha_total)] = f"=SUM({min_total_letter}{primeira}:{min_total_letter}{ultima})"
    ws[C(col_min_total, linha_total)].number_format = MOEDA

    ws[C(1, linha_frete)] = "FRETE"
    ws[C(1, linha_frete)].font = NEGRITO

    # frete somado ao total = frete do fornecedor com o MENOR total cheio
    # (comparação feita entre os totais da linha `linha_total`, coluna a
    # coluna) — assunção documentada, não é uma verdade absoluta quando os
    # fornecedores cobram frete diferente entre si.
    totais_letters = [get_column_letter(b["total"]) + str(linha_total) for b in blocos]
    fretes_letters = [get_column_letter(b["unit"]) + str(linha_frete) for b in blocos]
    expressao_frete = fretes_letters[-1]
    for tot, frete in list(zip(totais_letters, fretes_letters))[:-1]:
        expressao_frete = f'IF({tot}=MIN({",".join(totais_letters)}),{frete},{expressao_frete})'
    ws[C(col_min_unit, linha_frete)] = (
        f"={min_total_letter}{linha_total}+{expressao_frete}"
    )
    ws[C(col_min_unit, linha_frete)].number_format = MOEDA
    ws[C(col_min_unit, linha_frete)].font = NEGRITO
    ws[C(col_empresa, linha_frete)] = "PIX"

    # --- rodapé -------------------------------------------------------
    linha_necessidade = linha_frete + 2
    ws[C(1, linha_necessidade)] = "DATA DE NECESSIDADE"
    ws[C(1, linha_necessidade)].font = NEGRITO

    linha_local = linha_necessidade + 1
    ws[C(1, linha_local)] = "LOCAL ENTREGA (DESCARGA):"
    ws[C(1, linha_local)].font = NEGRITO

    linha_obs = linha_local + 1
    ws[C(1, linha_obs)] = "OBSERVAÇÕES GERAIS"
    ws[C(1, linha_obs)].font = NEGRITO

    # larguras básicas
    ws.column_dimensions["D"].width = 32
    for col in range(5, col_empresa + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15

    Path(caminho_saida).parent.mkdir(parents=True, exist_ok=True)
    wb.save(caminho_saida)
    return caminho_saida


if __name__ == "__main__":
    objeto = sys.argv[1] if len(sys.argv) > 1 else "Filamentos 3D"
    saida = sys.argv[2] if len(sys.argv) > 2 else "mapa_de_cotacao.xlsx"
    caminho = gerar(objeto, saida)
    print(f"Gerado: {caminho}")
