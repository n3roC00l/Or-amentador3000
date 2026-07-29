"""
Gera o pedido de cotação (RFQ) que sai da Cilla Tech Park para as empresas
cotarem — lista simples de material/cor/quantidade necessária, sem coluna
de preço (isso é o que a empresa vai preencher na resposta dela, não faz
sentido a gente já mandar um valor).

Uso:
    from relatorio_cotacao import gerar
    gerar([
        {"material": "PLA", "cor": "Azul", "quantidade_g": 3000, "observacao": ""},
        ...
    ], objeto="Reposição de estoque agosto/2026")
"""
import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

FONTE = "Arial"
NEGRITO = Font(name=FONTE, bold=True)
TITULO = Font(name=FONTE, bold=True, size=14)
CENTRO = Alignment(horizontal="center", vertical="center")


def gerar(
    itens: list[dict],
    objeto: str = "Cotação de filamentos",
    caminho_saida: str = "pedido_de_cotacao.xlsx",
) -> str:
    """
    `itens`: lista de dicts com chaves material, cor, quantidade_g e
    opcionalmente observacao. Levanta ValueError se a lista vier vazia —
    não gera planilha em branco silenciosamente.
    """
    if not itens:
        raise ValueError("nenhum filamento selecionado — escolha ao menos um item antes de gerar o pedido")

    wb = Workbook()
    ws = wb.active
    ws.title = "Pedido de Cotação"

    ws["A1"] = "PEDIDO DE COTAÇÃO — FILAMENTOS 3D"
    ws["A1"].font = TITULO

    ws["A2"] = "Objeto:"
    ws["A2"].font = NEGRITO
    ws["B2"] = objeto

    ws["A3"] = "Data:"
    ws["A3"].font = NEGRITO
    ws["B3"] = datetime.date.today()
    ws["B3"].number_format = "DD/MM/YYYY"

    linha_hdr = 5
    cabecalhos = ["Item", "Material", "Cor", "Quantidade Necessária (g)", "Observação"]
    for col, texto in enumerate(cabecalhos, start=1):
        celula = ws.cell(row=linha_hdr, column=col, value=texto)
        celula.font = NEGRITO
        celula.alignment = CENTRO

    for idx, item in enumerate(itens, start=1):
        r = linha_hdr + idx
        ws.cell(row=r, column=1, value=idx)
        ws.cell(row=r, column=2, value=item["material"])
        ws.cell(row=r, column=3, value=item["cor"])
        ws.cell(row=r, column=4, value=item["quantidade_g"])
        ws.cell(row=r, column=5, value=item.get("observacao", ""))

    larguras = {"A": 6, "B": 14, "C": 20, "D": 26, "E": 34}
    for col, largura in larguras.items():
        ws.column_dimensions[col].width = largura

    Path(caminho_saida).parent.mkdir(parents=True, exist_ok=True)
    wb.save(caminho_saida)
    return caminho_saida
